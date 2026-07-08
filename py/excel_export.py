"""Exportación a Excel del resumen de una CLUES (equivalente a utils_clues.r: crear_excel)."""

import numpy as np
import openpyxl
import pandas as pd

from py.utils_comunes import fecha_corte


def _valor_clues(fila: pd.DataFrame, columna: str):
    if fila.empty:
        return None
    valor = fila[columna].iloc[0]
    return None if pd.isna(valor) else valor


def _escribir_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int = 1):
    for j, nombre_col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=str(nombre_col))

    df_limpio = df.replace({np.nan: None})
    for i, (_, fila) in enumerate(df_limpio.iterrows()):
        for j, valor in enumerate(fila):
            if isinstance(valor, pd.Timestamp):
                valor = valor.to_pydatetime()
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=valor)


def crear_excel(
    clues_seleccionada: str,
    clues_info: pd.DataFrame,
    historico: pd.DataFrame,
    resumen: pd.DataFrame,
) -> openpyxl.Workbook:
    """Genera el libro con hoja 'resumen' (metadatos + tabla resumen) y 'productividad detalle' (histórico)."""
    fila_clues = clues_info[clues_info["clues_imb"] == clues_seleccionada]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "resumen"

    ws.cell(row=1, column=1, value=clues_seleccionada)
    ws.cell(row=2, column=1, value=_valor_clues(fila_clues, "nombre_de_la_unidad"))
    ws.cell(row=3, column=1, value=_valor_clues(fila_clues, "entidad"))
    ws.cell(row=4, column=1, value=_valor_clues(fila_clues, "nivel_atencion"))
    # El original escribe categoria_gerencial_ampliada en la fila 5 y luego lo sobreescribe
    # con la fecha de corte en la misma celda; se replica el mismo comportamiento.
    ws.cell(row=5, column=1, value=f"Fecha de corte: {fecha_corte.strftime('%d/%m/%Y')}")

    _escribir_dataframe(ws, resumen, start_row=8)

    ws_detalle = wb.create_sheet("productividad detalle")
    _escribir_dataframe(ws_detalle, historico, start_row=1)

    return wb
